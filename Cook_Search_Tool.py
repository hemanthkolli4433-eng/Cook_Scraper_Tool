import os
import time
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ==========================================================
# CONFIG
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "Cook_input.txt"
OUTPUT_FILE = BASE_DIR / "Cook_output.xlsx"
URL = "https://www.cookcountypropertyinfo.com/default.aspx"
COUNTY = "Cook"
STATE = "IL"

# Cloud GitHub headless is blocked/unstable for this website.
# For working automation, run this from a Windows self-hosted runner in visible Chrome.
HEADLESS = os.environ.get("HEADLESS", "0").strip() == "1"
WAIT_SECONDS = int(os.environ.get("WAIT_SECONDS", "120"))
MAX_RETRIES = int(os.environ.get("MAX_RETRIES", "3"))
PAUSE_AFTER_SEARCH = float(os.environ.get("PAUSE_AFTER_SEARCH", "5"))
USE_PROFILE = os.environ.get("COOK_USE_PROFILE", "1").strip() == "1"

SCREENSHOT_DIR = BASE_DIR / "error_screenshots"
HTML_DIR = BASE_DIR / "error_pages"
LOG_DIR = BASE_DIR / "logs"
PROFILE_DIR = BASE_DIR / "Chrome_Cook_Profile"
for folder in (SCREENSHOT_DIR, HTML_DIR, LOG_DIR):
    folder.mkdir(exist_ok=True)

LOG_FILE = LOG_DIR / f"run_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

HEADERS = [
    "Serial Number", "County", "State", "Input",
    "Parcel", "Property Address", "City", "Zip", "Township",
    "Mailing Address",
    "Assessed Value", "Estimated Value", "Lot Size", "Building Size",
    "Tax Year 1", "Tax Amount 1", "Tax Detail 1",
    "Tax Year 2", "Tax Amount 2", "Tax Detail 2",
    "Tax Year 3", "Tax Amount 3", "Tax Detail 3",
    "Tax Year 4", "Tax Amount 4", "Tax Detail 4"
]


def log(message: str) -> None:
    text = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}"
    print(text, flush=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(text + "\n")


def safe_name(value: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in value).strip("_")


def save_debug(driver, pin: str, prefix: str = "error") -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"{prefix}_{safe_name(pin)}_{stamp}"

    try:
        log(f"Current URL: {driver.current_url}")
        log(f"Page title: {driver.title}")
    except Exception:
        pass

    try:
        png_path = SCREENSHOT_DIR / f"{base}.png"
        driver.save_screenshot(str(png_path))
        log(f"Saved error screenshot: {png_path.name}")
    except Exception as e:
        log(f"Could not save screenshot for {pin}: {e}")

    try:
        html_path = HTML_DIR / f"{base}.html"
        html_path.write_text(driver.page_source, encoding="utf-8", errors="ignore")
        log(f"Saved error page HTML: {html_path.name}")
    except Exception as e:
        log(f"Could not save HTML for {pin}: {e}")


def create_driver() -> webdriver.Chrome:
    options = Options()

    if HEADLESS:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
    else:
        options.add_argument("--start-maximized")

    if USE_PROFILE:
        PROFILE_DIR.mkdir(exist_ok=True)
        options.add_argument(f"--user-data-dir={PROFILE_DIR}")
        options.add_argument("--profile-directory=Default")

    # Stability + reduce automation detection.
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-notifications")
    options.add_argument("--remote-allow-origins=*")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(WAIT_SECONDS)
    try:
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    except Exception:
        pass
    return driver


def get_or_create_workbook():
    if OUTPUT_FILE.exists():
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        if ws.max_row < 1 or ws.cell(row=1, column=1).value != HEADERS[0]:
            ws.insert_rows(1)
            for col, header in enumerate(HEADERS, 1):
                ws.cell(row=1, column=col).value = header
            wb.save(OUTPUT_FILE)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cook Output"
        ws.append(HEADERS)
        wb.save(OUTPUT_FILE)
    return wb, ws


def already_saved_inputs(ws):
    saved = set()
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=4).value  # Column D = Input
        if value:
            saved.add(str(value).strip())
    return saved


def read_pins():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")
    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        pins = [line.strip() for line in f if line.strip()]

    unique = []
    seen = set()
    for pin in pins:
        if pin not in seen:
            unique.append(pin)
            seen.add(pin)
    return unique


def wait_for_ready(driver):
    WebDriverWait(driver, WAIT_SECONDS).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )


def clear_and_type(element, value: str) -> None:
    element.click()
    element.send_keys(Keys.CONTROL, "a")
    element.send_keys(Keys.BACKSPACE)
    element.send_keys(value)


def get_text(driver, xpath: str) -> str:
    try:
        return driver.find_element(By.XPATH, xpath).text.strip()
    except Exception:
        return ""


def wait_for_result_or_block(driver, wait):
    def condition(d):
        src = d.page_source.lower()
        if d.find_elements(By.ID, "ContentPlaceHolder1_lblResultTitle"):
            return "result"
        if "captcha" in src or "access denied" in src or "verify" in src or "blocked" in src:
            return "blocked"
        if "pinbox1" in src and "lblresulttitle" not in src:
            return False
        return False

    return wait.until(condition)


def input_pin_and_search(driver, wait, pin: str) -> None:
    pin_parts = pin.split("-")
    if len(pin_parts) != 5:
        raise ValueError(f"Invalid PIN format: {pin}. Expected format like 06-27-300-027-0000")

    driver.get(URL)
    wait_for_ready(driver)

    fields = ["pinBox1", "pinBox2", "pinBox3", "pinBox4", "pinBox5"]
    for field_id, value in zip(fields, pin_parts):
        element = wait.until(EC.element_to_be_clickable((By.ID, field_id)))
        clear_and_type(element, value)
        time.sleep(0.15)

    search_btn = wait.until(EC.element_to_be_clickable((By.ID, "ContentPlaceHolder1_PINAddressSearch_btnSearch")))
    try:
        search_btn.click()
    except Exception:
        driver.execute_script("arguments[0].click();", search_btn)

    status = wait_for_result_or_block(driver, wait)
    if status == "blocked":
        raise TimeoutException("Website looks blocked/captcha/verify page appeared.")

    time.sleep(PAUSE_AFTER_SEARCH)


def scrape_current_result(driver, wait, pin: str):
    parcel = get_text(driver, '//*[@id="ContentPlaceHolder1_lblResultTitle"]')
    prop_address = get_text(driver, '//*[@id="ContentPlaceHolder1_PropertyInfo_propertyAddress"]')
    city = get_text(driver, '//*[@id="ContentPlaceHolder1_PropertyInfo_propertyCity"]')
    zip_code = get_text(driver, '//*[@id="ContentPlaceHolder1_PropertyInfo_propertyZip"]')
    township = get_text(driver, '//*[@id="ContentPlaceHolder1_PropertyInfo_propertyTownship"]')

    try:
        mailing_table = driver.find_element(
            By.XPATH,
            '//*[@id="ContentPlaceHolder1_success"]/div/div[2]/div[2]/table[2]/tbody'
        )
        lines = []
        for tr in mailing_table.find_elements(By.TAG_NAME, "tr"):
            parts = [td.text.strip() for td in tr.find_elements(By.TAG_NAME, "td") if td.text.strip()]
            if parts:
                lines.append(" ".join(parts))
        mailing_address = " | ".join(lines)
    except Exception:
        mailing_address = ""

    assessed_value = get_text(driver, '//*[@id="ContentPlaceHolder1_TaxYearInfo_propertyAssessedValue"]')
    estimated_value = get_text(driver, '//*[@id="ContentPlaceHolder1_TaxYearInfo_propertyEstimatedValue"]')
    lot_size = get_text(driver, '//*[@id="ContentPlaceHolder1_TaxYearInfo_propertyLotSize"]')
    building_size = get_text(driver, '//*[@id="ContentPlaceHolder1_TaxYearInfo_propertyBuildingSize"]')

    try:
        tax_section = wait.until(
            EC.presence_of_element_located((By.XPATH, '//div[contains(text(), "TAX BILLED AMOUNTS")]'))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tax_section)
        time.sleep(1)
    except Exception:
        log(f"Tax section scroll failed for {pin}. Continuing.")

    tax_data = []
    try:
        tax_table = driver.find_element(
            By.XPATH,
            '//table[contains(@class,"property") and contains(@class,"categorybottomspace2")]/tbody'
        )
        tax_rows = tax_table.find_elements(By.TAG_NAME, "tr")[:4]
        for row in tax_rows:
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) >= 3:
                tax_data.extend([tds[0].text.strip(), tds[1].text.strip(), tds[2].text.strip()])
    except Exception as e:
        log(f"Failed to scrape tax table for {pin}: {e}")

    while len(tax_data) < 12:
        tax_data.extend(["", "", ""])

    return [
        COUNTY, STATE, pin,
        parcel, prop_address, city, zip_code, township,
        mailing_address,
        assessed_value, estimated_value, lot_size, building_size
    ] + tax_data


def process_one_pin(driver, pin: str):
    wait = WebDriverWait(driver, WAIT_SECONDS)
    last_error = None

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            log(f"Processing: {pin} | attempt {attempt}/{MAX_RETRIES}")
            input_pin_and_search(driver, wait, pin)
            row_data_without_serial = scrape_current_result(driver, wait, pin)

            parcel = row_data_without_serial[3]
            prop_address = row_data_without_serial[4]
            if not parcel and not prop_address:
                raise TimeoutException("Result page loaded, but Parcel and Property Address are blank.")

            return row_data_without_serial

        except Exception as e:
            last_error = e
            log(f"Error with {pin} on attempt {attempt}: {repr(e)}")
            save_debug(driver, pin, prefix=f"attempt_{attempt}")
            try:
                driver.get(URL)
                wait_for_ready(driver)
                time.sleep(2)
            except Exception:
                pass

    raise last_error


def main():
    log("Cook County search tool started")
    log(f"Base folder: {BASE_DIR}")
    log(f"Headless mode: {HEADLESS}")
    log(f"Use Chrome profile: {USE_PROFILE}")
    log(f"Wait seconds: {WAIT_SECONDS}")
    log(f"Max retries: {MAX_RETRIES}")

    pins = read_pins()
    log(f"Total input PINs: {len(pins)}")

    wb, ws = get_or_create_workbook()
    saved_inputs = already_saved_inputs(ws)
    log(f"Already saved PINs in output: {len(saved_inputs)}")

    driver = None
    try:
        driver = create_driver()

        for pin in pins:
            if pin in saved_inputs:
                log(f"Skipping already saved PIN: {pin}")
                continue

            try:
                row_data_without_serial = process_one_pin(driver, pin)
                serial = ws.max_row
                ws.append([serial] + row_data_without_serial)
                wb.save(OUTPUT_FILE)
                saved_inputs.add(pin)
                log(f"Saved: {pin}")

            except WebDriverException as e:
                log(f"WebDriver error after retries for {pin}: {repr(e)}")
                log(traceback.format_exc())
                try:
                    if driver:
                        driver.quit()
                except Exception:
                    pass
                driver = create_driver()
                continue

            except Exception as e:
                log(f"Final failed after all retries for {pin}: {repr(e)}")
                log(traceback.format_exc())
                continue

    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass
        try:
            wb.save(OUTPUT_FILE)
        except Exception:
            pass
        log(f"Finished. Output file: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
